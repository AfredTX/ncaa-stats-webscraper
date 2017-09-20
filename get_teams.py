#!/usr/bin/env python3
"""
Python webscrapping module.

Webscrap teamrankings.com website to obtain a list of the top 68 basketball
teams and url references to use for webscrapping other websites for team stats.

For each of the 68 teams, webscrap teamrankings.com and kenpom.com for lists
of 10 key stats, then write those stats to an existing Excel spreadsheet.
"""
import re
import time
from difflib import get_close_matches
from itertools import islice

import pandas
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.compat import range
from progressbar import Bar, Percentage, ProgressBar

# create ProgressBar objects
p1 = ProgressBar(widgets=[Percentage(), Bar()], maxval=300)
p2 = ProgressBar(widgets=[Percentage(), Bar()], maxval=300)
p3 = ProgressBar(widgets=[Percentage(), Bar()], maxval=300)
p4 = ProgressBar(widgets=[Percentage(), Bar()], maxval=300)
p5 = ProgressBar(widgets=[Percentage(), Bar()], maxval=300)
p6 = ProgressBar(widgets=[Percentage(), Bar()], maxval=300)
p7 = ProgressBar(widgets=[Percentage(), Bar()], maxval=300)
p8 = ProgressBar(widgets=[Percentage(), Bar()], maxval=300)

start = time.time()
# assign url to variable and use requests to get html
print('Getting a list of all teams.')
url = 'https://www.teamrankings.com/ncb/'
teams_page = requests.get(url)
teams = teams_page.content

# parse html with BeautifulSoup
soup = BeautifulSoup(teams, 'html.parser')

# scrap all team names and add to a dictionary
full_teams_list = {}
lst = soup.find_all('div', {'class': 'table-team-logo-text'})

for item in p1(lst):
    team_name = item.find('a').text
    # scrap url reference for team name (e.g., 'gonzaga-bulldogs') and add to
    # dictionary
    url_ref = item.find('a')['href'][22:]
    full_teams_list[team_name] = url_ref
print('Done!')


def take(n, iterable):
    """Return first n items of the iterable as a list."""
    return list(islice(iterable, n))


# grab top 68 teams and create list of teams and list of url references
print('Cutting that list to the top 68 teams')
top_68 = take(68, full_teams_list.items())
teams_list = []
url_ref_list = []
for tup in p2(top_68):
    teams_list.append(tup[0])
    url_ref_list.append(tup[1])
print('Done!')

# build two lists of urls to scrap stats for each team
print('Building two lists of urls to parse')
stat_url_list = []
sos_url_list = []
for url in p3(url_ref_list):
    stat_url_list.append('https://www.teamrankings.com/ncaa-basketball/team/' +
                         url + '/stats')
    sos_url_list.append('https://www.teamrankings.com/ncaa-basketball/team/' +
                        url + '/rankings')
print('Done!')

print('Getting stats from teamrankings.com')
# create empty lists to hold stats
eFG = []
TO = []
OR = []
FT = []
deFG = []
dTO = []
DR = []
dFT = []
five = []
# loop through the first url list for each team to get the stats and add to the
# empty lists
print('First Four')
for stat in p4(stat_url_list):
    stats_page = requests.get(stat)
    stats = stats_page.content

    soup1 = BeautifulSoup(stats, 'html.parser')

    numbers = soup1.find_all('td', {'class': 'nowrap'})
    # temporary list
    s = []
    for number in numbers:
        # this will add all the stats from the url, even those we don't need
        s.append(number.text)
    # use list indexing and slicing to add only the stats we want
    eFG.append(s[25][:5])
    TO.append(s[125][:5])
    OR.append(s[97][:5])
    FT.append(s[29][:5])
    deFG.append(s[27][:5])
    dTO.append(s[127][:5])
    DR.append(s[101][:5])
    dFT.append(s[31][:5])

# same thing but for the second url list
print('Strength of Schedule')
for sos in p5(sos_url_list):
    rankings_page = requests.get(sos)
    rankings = rankings_page.content

    soup2 = BeautifulSoup(rankings, 'html.parser')

    numbers2 = soup2.find_all('td', {'class': 'text-right'})

    s1 = []
    for number2 in numbers2:
        s1.append(number2.text)
    five.append(s1[15])
print('Done!')

# assign url to variable and use requests to get html
print('Getting stats from kenpom.com')
url1 = 'https://kenpom.com/'
kp_page = requests.get(url1)
kp = kp_page.content

# parse html with BeautifulSoup
soup3 = BeautifulSoup(kp, 'html.parser')

# create lists of the team names and stats (again, all teams and stats)
left = soup3.find_all('td', {'class': 'td-left'})
names = soup3.find_all('a', {'href': re.compile('team\.php\?team=.+')})
s2 = []
n = []
print('Stats')
for number3 in p6(left):
    s2.append(number3.text)
print('Teams')
for name in p7(names):
    n.append(name.text)


def swap(old_team, new_team):
    """Replace old_team with new_team for list n."""
    i = n.index(old_team)
    n.remove(old_team)
    n.insert(i, new_team)


# replace team names so they match
swap('North Carolina', 'N Carolina')
swap('West Virginia', 'W Virginia')
swap('Iowa St.', 'Iowa State')
swap('SMU', 'S Methodist')
swap('Saint Mary\'s', 'St Marys')
swap('South Carolina', 'S Carolina')
swap('Miami FL', 'Miami (FL)')
swap('TCU', 'TX Christian')
swap('Virginia Tech', 'VA Tech')
swap('Middle Tennessee', 'Middle Tenn')
swap('UNC Wilmington', 'NC-Wilmgton')
swap('Ohio St.', 'Ohio State')

# use list slicing to get only the stats we need and add to two new lists
adjo = s2[0::8]
adjd = s2[1::8]
# create a third list to hold the difference of lists one and two
diff = []
for o, d in zip(adjo, adjd):
    diff.append(round(float(o) - float(d), 1))

print('Done!')

# convert the four lists into a DataFrame object
df2 = pandas.DataFrame([n, adjo, adjd, diff]).T
# set the team name as the DataFrame index
df2 = df2.set_index(0)

# convert the DataFrame to a dictionary, splitting the team names and stats
# into separate dictionaries
df3 = pandas.DataFrame.to_dict(df2, orient='split')


def getStats(team):
    """Return list of stats for team if that team is found or return 0.0."""
    if team in df3['index']:
        index = df3['index'].index(team)
        return df3['data'][index]
    elif len(get_close_matches(team, df3['index'], cutoff=0.92)) > 0:
        b = get_close_matches(team, df3['index'], cutoff=0.92)[0]
        index = df3['index'].index(b)
        return df3['data'][index]
    else:
        return [0.0, 0.0, 0.0]


# create new lists to hold the stats for the top 68 teams only
_adjo = []
_adjd = []
_diff = []
# use the getStats function to match team names from both websites and get the
# stats from kenpom.com for that team
print('Getting the right stats for the right teams')
for team in p8(teams_list):
    _adjo.append(getStats(team)[0])
    _adjd.append(getStats(team)[1])
    _diff.append(getStats(team)[2])

# load workbook object from Excel spreadsheet
print('Adding to NCAA Bracket Spreadsheet')
wb = load_workbook(filename='NCAA Bracket Spreadsheet.xlsx')
sheet = wb.get_sheet_by_name('Provided Ranking')


def writeToExcel(list, col):
    """Write data to the Excel spreadsheet."""
    counter = 0
    for rowNum in range(3, 71):
        sheet.cell(row=rowNum, column=col).value = list[counter]
        counter += 1


# use writeToExcel function to stats from all 13 lists to the Excel spreadsheet
pbar = ProgressBar(widgets=[Percentage(), Bar()], maxval=300).start()
writeToExcel(teams_list, 2)
pbar.update(8)
writeToExcel(eFG, 3)
pbar.update(7)
writeToExcel(TO, 4)
pbar.update(8)
writeToExcel(OR, 5)
pbar.update(7)
writeToExcel(FT, 6)
pbar.update(8)
writeToExcel(deFG, 7)
pbar.update(7)
writeToExcel(dTO, 8)
pbar.update(8)
writeToExcel(DR, 9)
pbar.update(7)
writeToExcel(dFT, 10)
pbar.update(8)
writeToExcel(five, 14)
pbar.update(7)
writeToExcel(_adjo, 11)
pbar.update(8)
writeToExcel(_adjd, 12)
pbar.update(7)
writeToExcel(_diff, 13)
pbar.update(8)
pbar.finish()

# save copy of Excel spreadsheet
wb.save(filename='NCAA Bracket Spreadsheet-copy.xlsx')
print('Done!')

end = time.time()
total_time = round((end - start) / 60, 2)
print('All steps complete.\nScript completed in %s minutes.\nGood luck!' %
      total_time)
